from fastapi import FastAPI, Depends, HTTPException, Header, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import os
import json
import time
import threading
import urllib.request
import urllib.parse
from datetime import datetime

from database import get_db, init_db

# ── Папка для загрузок ────────────────────────────────────────────────────────
UPLOADS_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

# ── Глобальный трекер процессов ──────────────────────────────────────────────
# _state_lock защищает _PROCESS_STATUS и _history_import_state от inconsistent
# read/write между sync loop thread, history import bg thread и HTTP-endpoint'ами.
_state_lock = threading.Lock()

_PROCESS_STATUS = {
    "kaspi_sync": {
        "status": "starting",   # starting | running | idle | error
        "last_run": None,
        "next_run": None,
        "last_result": None,
        "last_error": None,
        "cycle_count": 0,
    },
    "server_start": datetime.utcnow().isoformat(),
}
APP_START = time.monotonic()

from helpers import save_upload as _save_upload  # noqa: E402
import crud
import kaspi as kaspi_module




from helpers import decode_kaspi_order_id as _decode_kaspi_order_id  # noqa: E402


from helpers import parse_order_date as _parse_order_date, filter_orders_by_date as _filter_orders_by_date  # noqa: E402

import secrets
from fastapi import Request, Cookie
from fastapi.responses import JSONResponse, RedirectResponse
from starlette.middleware.base import BaseHTTPMiddleware

from helpers import (  # noqa: E402
    make_session_token as _make_session_token,
    SESSION_TOKEN as _SESSION_TOKEN,
)

try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    limiter = Limiter(key_func=get_remote_address, default_limits=["120/minute"])
    _slowapi_ok = True
except ImportError:
    _slowapi_ok = False

app = FastAPI(title="Lunary OS", version="1.0")

# ── Роутеры (рефакторинг api.py из монолита) ──
from routers.analytics import router as analytics_router  # noqa: E402
from routers.uploads import router as uploads_router  # noqa: E402
from routers.pricelist import router as pricelist_router  # noqa: E402
from routers.merge import router as merge_router  # noqa: E402
from routers.store import router as store_router  # noqa: E402
from routers.products import router as products_router  # noqa: E402
from routers.kaspi import router as kaspi_router  # noqa: E402
from routers.settings import router as settings_router  # noqa: E402
from routers.shop_orders import router as shop_orders_router  # noqa: E402
from routers.system import router as system_router  # noqa: E402
from routers.review import router as review_router  # noqa: E402
from routers.admin import router as admin_router  # noqa: E402
from routers.ai import router as ai_router  # noqa: E402
from routers.seo import router as seo_router  # noqa: E402
from routers.aliases import router as aliases_router  # noqa: E402
from routers.audit import router as audit_router  # noqa: E402
app.include_router(analytics_router)
app.include_router(uploads_router)
app.include_router(pricelist_router)
app.include_router(merge_router)
app.include_router(store_router)
app.include_router(products_router)
app.include_router(kaspi_router)
app.include_router(settings_router)
app.include_router(shop_orders_router)
app.include_router(system_router)
app.include_router(review_router)
app.include_router(admin_router)
app.include_router(ai_router)
app.include_router(seo_router)
app.include_router(aliases_router)
app.include_router(audit_router)


if _slowapi_ok:
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


# ─── Auth middleware ──────────────────────────────────────────
_PUBLIC_PATHS = {"/login", "/api/auth/login", "/api/auth/logout",
                 "/auth/google", "/auth/google/callback",
                 "/shop", "/", "/about", "/api/settings"}
_PUBLIC_PREFIXES = ("/static/", "/api/store/", "/api/shop/", "/shop/product/")
_ADMIN_PATHS = ("/admin", "/kaspi", "/analytics", "/history", "/scanner",
                "/api/kaspi", "/api/analytics", "/api/products", "/api/movements",
                "/api/history", "/api/admin", "/api/purchases", "/api/alerts",
                "/merge", "/review", "/import", "/pricelist", "/uploads", "/api/merge", "/api/import-price-list",
                "/api/kaspi/import", "/api/reset-products", "/api/fill-articles", "/api/pricelist", "/api/uploads")


from helpers import (  # noqa: E402
    get_user_from_session as _get_user_from_session,
    is_staff as _is_staff,
    is_admin as _is_admin,
)

# Пути доступные manager (не только admin)
_MANAGER_ALLOWED_PREFIXES = (
    "/admin", "/kaspi", "/analytics", "/history", "/scanner",
    "/api/kaspi", "/api/analytics", "/api/products", "/api/movements",
    "/api/history", "/api/purchases", "/api/alerts",
    "/merge", "/review", "/import", "/pricelist", "/uploads",
    "/api/merge", "/api/import-price-list", "/api/pricelist", "/api/uploads",
)
# Пути только для admin (опасные операции)
_ADMIN_ONLY_PREFIXES = (
    "/api/admin/settings", "/api/admin/users", "/api/reset-products",
    "/api/fill-articles", "/api/admin/fill", "/admin/settings", "/admin/theme",
)


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        method = request.method

        # Публичные пути — без авторизации
        if path in _PUBLIC_PATHS or any(path.startswith(p) for p in _PUBLIC_PREFIXES):
            return await call_next(request)

        user = _get_user_from_session(request)

        # Только-admin пути (настройки, сброс, пользователи)
        is_admin_only = any(path.startswith(p) for p in _ADMIN_ONLY_PREFIXES)
        if is_admin_only:
            if not _is_admin(user):
                if path.startswith("/api/"):
                    return JSONResponse({"detail": "Forbidden"}, status_code=403)
                return RedirectResponse(f"/login?next={path}", status_code=302)

        # Удаление — только admin
        if method == "DELETE" and any(path.startswith(p) for p in ("/api/products", "/api/kaspi", "/api/movements")):
            if not _is_admin(user):
                return JSONResponse({"detail": "Только администратор может удалять"}, status_code=403)

        # Админские пути — admin или manager
        is_admin_path = any(path.startswith(p) for p in _ADMIN_PATHS)
        if is_admin_path:
            if not _is_staff(user):
                if path.startswith("/api/"):
                    return JSONResponse({"detail": "Forbidden"}, status_code=403)
                return RedirectResponse(f"/login?next={path}", status_code=302)

        # Пути требующие авторизации (оформление заказа и т.д.)
        _AUTH_REQUIRED = ("/api/orders",)
        if any(path.startswith(p) for p in _AUTH_REQUIRED):
            if not user:
                return JSONResponse({"detail": "Unauthorized"}, status_code=401)

        return await call_next(request)

app.add_middleware(AuthMiddleware)
app.mount("/static", StaticFiles(directory="static"), name="static")


# ─── Pydantic схемы ──────────────────────────────────────────
# ProductCreate/Update/StockAdjust переехали в routers/products.py



from helpers import get_integration as _get_integration  # noqa: E402


# ─── Startup ─────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    init_db()
    _auto_import_if_empty()
    _start_kaspi_sync_loop()


def _start_kaspi_sync_loop():
    """Фоновая синхронизация заказов Kaspi каждые 5 минут"""

    from database import KaspiOrder, SessionLocal as SL

    STATES = ["NEW", "APPROVED", "PICKUP", "DELIVERY", "KASPI_DELIVERY", "ARCHIVE", "CANCELLED", "SIGN_REQUIRED"]

    _backfill_fetched_this_cycle = [0]  # только для дозагрузки старых заказов
    MAX_BACKFILL_PER_CYCLE = 30         # не более 30 старых заказов за цикл

    def _fetch_entries(oid, is_new=False):
        """Загружает состав заказа. Новые заказы — всегда, старые — до 30 за цикл."""
        if not is_new and _backfill_fetched_this_cycle[0] >= MAX_BACKFILL_PER_CYCLE:
            return []
        try:
            result = kaspi_module.get_order_entries(oid) or []
            print(f"[sync entries] oid={oid} is_new={is_new} got={len(result)}", flush=True)
            if not is_new:
                _backfill_fetched_this_cycle[0] += 1
                time.sleep(0.2)
            return result
        except Exception as e:
            print(f"[sync entries] oid={oid} ОШИБКА: {e}", flush=True)
            return []

    def _update_entries_fields(row, entries):
        """Обновляет entries + product_name/sku/quantity"""
        row.entries = json.dumps(entries, ensure_ascii=False)
        if entries and not row.product_name:
            row.product_name = entries[0].get("name")
            row.sku = entries[0].get("merchantSku", "")
            row.quantity = sum(e.get("qty", 1) for e in entries if isinstance(e, dict))

    def sync():
        while True:
            cycle_start = time.monotonic()
            with _state_lock:
                _PROCESS_STATUS["kaspi_sync"]["status"] = "running"
                _PROCESS_STATUS["kaspi_sync"]["last_run"] = datetime.utcnow().isoformat()
                _PROCESS_STATUS["kaspi_sync"]["cycle_count"] += 1
            db = SL()
            try:
                from database import SyncLog
                from datetime import timezone, timedelta
                tz_kz = timezone(timedelta(hours=5))
                sync_start = datetime.utcnow()

                all_orders = []
                for state in STATES:
                    result = kaspi_module.get_kaspi_orders(state=state, size=100)
                    if result.get("orders"):
                        all_orders.extend(result["orders"])

                # Дедупликация: один заказ может прийти из нескольких state-запросов
                seen_ids = {}
                for o in all_orders:
                    oid = _decode_kaspi_order_id(o["id"])
                    if oid not in seen_ids:
                        seen_ids[oid] = o
                all_orders = list(seen_ids.values())

                added = 0
                updated_count = 0
                returns_count = 0
                deducted_count = 0
                new_orders = []
                _backfill_fetched_this_cycle[0] = 0  # сброс счётчика

                # Один раз строим индекс SKU→product_id, передаём во все вызовы
                # deduct/return. Избавляет от N full-scan'ов таблицы products
                # при обработке каждого заказа.
                from helpers import build_sku_index
                sku_index = build_sku_index(db)

                for o in all_orders:
                    raw_date = o.get("date", "")
                    try:
                        d = _parse_order_date(raw_date)
                        order_date = d.strftime("%d.%m.%Y") if d else str(raw_date)
                    except Exception:
                        order_date = str(raw_date)

                    oid = _decode_kaspi_order_id(o["id"])
                    addr_obj = o.get("deliveryAddress")
                    address = addr_obj.get("formattedAddress", "") if isinstance(addr_obj, dict) else str(addr_obj or "")

                    existing = db.query(KaspiOrder).filter(KaspiOrder.order_id == oid).first()
                    if not existing and oid != str(o["id"]):
                        existing = db.query(KaspiOrder).filter(KaspiOrder.order_id == str(o["id"])).first()
                        if existing:
                            existing.order_id = oid

                    if existing:
                        old_state = existing.state
                        new_state = o.get("state", existing.state)

                        # Остатки: списываем / возвращаем при смене статуса
                        if new_state in DEDUCT_STATES and old_state not in DEDUCT_STATES and not existing.stock_deducted:
                            _deduct_stock_for_order(existing, db, sku_index)
                            existing.stock_deducted = 1
                            deducted_count += 1
                        elif new_state in CANCEL_STATES and old_state not in CANCEL_STATES and existing.stock_deducted:
                            _return_stock_for_order(existing, db, sku_index)
                            existing.stock_deducted = 0
                            returns_count += 1

                        if new_state != old_state:
                            updated_count += 1

                        # Всегда обновляем все поля из свежего ответа Kaspi
                        existing.state = new_state
                        existing.last_synced_at = datetime.utcnow()
                        existing.total = int(o.get("total", existing.total or 0))
                        existing.customer = o.get("customer", existing.customer)
                        existing.delivery_method = o.get("deliveryMode", existing.delivery_method)
                        existing.payment_method = o.get("paymentMode", existing.payment_method)
                        if address:
                            existing.address = address
                        if new_state == "ARCHIVE" and old_state != "ARCHIVE" and not existing.status_date:
                            existing.status_date = datetime.now(tz=tz_kz).strftime("%d.%m.%Y")

                        # Грузим состав если пустой — только для активных заказов
                        # Kaspi не отдаёт entries для архивных/завершённых заказов
                        ACTIVE_FOR_ENTRIES = {"NEW", "APPROVED", "DELIVERY", "KASPI_DELIVERY", "PICKUP", "SIGN_REQUIRED"}
                        if (not existing.entries or existing.entries in ("[]", "")) and new_state in ACTIVE_FOR_ENTRIES:
                            entries = _fetch_entries(oid, is_new=False)
                            if entries:
                                _update_entries_fields(existing, entries)
                                o["entries"] = entries
                    else:
                        # Новый заказ — грузим состав всегда без лимита
                        entries = _fetch_entries(oid, is_new=True)
                        product_name, sku, quantity = None, None, None
                        if entries:
                            product_name = entries[0].get("name")
                            sku = entries[0].get("merchantSku") or (entries[0].get("sku", "") if isinstance(entries[0], dict) else None)
                            quantity = sum(e.get("qty", e.get("quantity", 1)) for e in entries if isinstance(e, dict))

                        new_ko = KaspiOrder(
                            order_id=oid,
                            state=o.get("state", ""),
                            total=int(o.get("total", 0)),
                            customer=o.get("customer", ""),
                            entries=json.dumps(entries, ensure_ascii=False),
                            order_date=order_date,
                            product_name=product_name,
                            sku=sku,
                            quantity=quantity,
                            delivery_method=o.get("deliveryMode", ""),
                            payment_method=o.get("paymentMode", ""),
                            address=address,
                            source="kaspi_api",
                            last_synced_at=datetime.utcnow(),
                        )
                        db.add(new_ko)
                        added += 1
                        o["id"] = oid
                        o["entries"] = entries
                        new_orders.append(o)

                # Записываем лог синхронизации
                db.add(SyncLog(
                    synced_at=sync_start.replace(tzinfo=None),
                    total_found=len(all_orders),
                    added=added,
                    updated=updated_count,
                    returns=returns_count,
                    deducted=deducted_count,
                ))
                db.commit()
                # Удаляем записи старше 1 дня
                cutoff = datetime.utcnow() - timedelta(days=1)
                db.query(SyncLog).filter(SyncLog.synced_at < cutoff).delete(synchronize_session=False)
                db.commit()

                notify_states = {"NEW", "PICKUP", "KASPI_DELIVERY", "DELIVERY"}
                for o in new_orders:
                    if o.get("state") in notify_states:
                        _send_tg_notification(_format_order_notification(o))

                with _state_lock:
                    _PROCESS_STATUS["kaspi_sync"]["last_result"] = f"+{added} новых, обновлено {updated_count}"
                    _PROCESS_STATUS["kaspi_sync"]["last_error"] = None
                if added or returns_count:
                    print(f"✅ Kaspi sync: +{added} новых, обновлено {updated_count}, возвратов {returns_count}")
            except Exception as e:
                print(f"⚠️ Kaspi sync error: {e}")
                with _state_lock:
                    _PROCESS_STATUS["kaspi_sync"]["status"] = "error"
                    _PROCESS_STATUS["kaspi_sync"]["last_error"] = str(e)[:300]
                try:
                    from database import SyncLog
                    db.add(SyncLog(synced_at=datetime.utcnow(), error=str(e)[:500]))
                    db.commit()
                except Exception:
                    pass
            finally:
                db.close()
            # Ждём ровно 5 минут от начала цикла (не от конца)
            elapsed = time.monotonic() - cycle_start
            sleep_sec = max(0, 300 - elapsed)
            with _state_lock:
                _PROCESS_STATUS["kaspi_sync"]["status"] = "idle"
                _PROCESS_STATUS["kaspi_sync"]["next_run"] = (datetime.utcnow().timestamp() + sleep_sec)
            time.sleep(sleep_sec)

    t = threading.Thread(target=sync, daemon=True)
    t.start()


def _auto_import_if_empty():
    """Если база пустая — автоматически импортируем товары из export_products.json"""
    import os
    from database import Product as _P, SessionLocal
    db = SessionLocal()
    try:
        count = db.query(_P).count()
        if count > 0:
            return
        path = os.path.join(os.path.dirname(__file__), '_archive', 'export_products.json')
        if not os.path.exists(path):
            return
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        added = 0
        for p in data:
            sku = p['sku'].upper()
            if False:  # sku field removed
                continue
            new_p = crud.create_product(
                name=p['name'], sku=sku, db=db,
                barcode=p.get('barcode'), category=p.get('category', 'Общее'),
                unit=p.get('unit', 'шт'), min_stock=p.get('min_stock', 5),
                brand=p.get('brand'), price=p.get('price')
            )
            if p.get('stock', 0) > 0:
                crud.set_initial_stock(new_p.id, p['stock'], db)
            added += 1
        print(f"✅ Автоимпорт: {added} товаров загружено из export_products.json")
    except Exception as e:
        print(f"⚠️ Автоимпорт ошибка: {e}")
    finally:
        db.close()


def _seed_kaspi_from_xml():
    """
    Первый старт на Railway: если kaspi_orders пустая — загрузить
    исторические заказы из XML. После этого синк с API дополняет новыми.
    """
    import os
    from database import KaspiOrder, SessionLocal
    db = SessionLocal()
    try:
        count = db.query(KaspiOrder).count()
        if count > 0:
            print(f"ℹ️ kaspi_orders уже содержит {count} заказов — XML пропущен")
            return
        xml_path = os.path.join(os.path.dirname(__file__), 'lunary_all_orders (1).xml')
        if not os.path.exists(xml_path):
            print("⚠️ XML файл не найден — исторические заказы не загружены")
            return
        import sys as _sys
        _sys.path.insert(0, os.path.join(os.path.dirname(__file__), '_archive'))
        import import_xml
        orders = import_xml.parse_orders(xml_path)
        inserted, updated = import_xml.upsert_orders(orders)
        print(f"✅ Исторические заказы загружены: {inserted} добавлено, {updated} обновлено")
    except Exception as e:
        print(f"⚠️ Ошибка загрузки исторических заказов: {e}")
    finally:
        db.close()




# ─── Auth ────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    password: str

@app.get("/login", response_class=HTMLResponse)
def login_page():
    with open("static/login.html", encoding="utf-8") as f:
        return f.read()

@app.post("/api/auth/login")
def auth_login(data: LoginRequest):
    password = os.getenv("ADMIN_PASSWORD", "")
    if not password:
        # Пароль не задан — вход ЗАПРЕЩЁН, требуется настройка
        raise HTTPException(status_code=503, detail="ADMIN_PASSWORD не задан — настройте переменную окружения на сервере")
    elif data.password == password:
        resp = JSONResponse({"ok": True})
    else:
        raise HTTPException(status_code=401, detail="Неверный пароль")
    resp.set_cookie(
        "lunary_session", _SESSION_TOKEN,
        httponly=True, samesite="lax", max_age=60 * 60 * 24 * 30  # 30 дней
    )
    return resp

@app.post("/api/auth/logout")
def auth_logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("lunary_session")
    return resp


class RegisterRequest(BaseModel):
    name: str
    email: str
    password: str


class EmailLoginRequest(BaseModel):
    email: str
    password: str


def _make_user_session(user) -> str:
    import hashlib
    secret = os.getenv("ADMIN_PASSWORD", "lunary-secret")
    return f"{user.id}_{hashlib.sha256(f'user-{user.id}-{user.email}-{secret}'.encode()).hexdigest()}"


@app.post("/api/auth/register")
def auth_register(data: RegisterRequest, db: Session = Depends(get_db)):
    import hashlib
    from database import User as UserModel
    if db.query(UserModel).filter(UserModel.email == data.email.lower()).first():
        raise HTTPException(status_code=409, detail="Пользователь с таким email уже существует")
    pw_hash = hashlib.sha256(data.password.encode()).hexdigest()
    _ADMIN_EMAILS = {e.strip().lower() for e in os.getenv("ADMIN_EMAILS", "aiseckiy@gmail.com").split(",")}
    role = "admin" if data.email.lower() in _ADMIN_EMAILS else "customer"
    user = UserModel(email=data.email.lower(), name=data.name, password_hash=pw_hash, role=role)
    db.add(user)
    db.commit()
    db.refresh(user)
    resp = JSONResponse({"ok": True, "name": user.name, "role": user.role})
    resp.set_cookie("lunary_session", _make_user_session(user), httponly=True, samesite="lax", max_age=60*60*24*30)
    return resp


@app.post("/api/auth/email-login")
def auth_email_login(data: EmailLoginRequest, db: Session = Depends(get_db)):
    import hashlib
    from database import User as UserModel
    user = db.query(UserModel).filter(UserModel.email == data.email.lower()).first()
    if not user or not user.password_hash:
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
    pw_hash = hashlib.sha256(data.password.encode()).hexdigest()
    if user.password_hash != pw_hash:
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
    _ADMIN_EMAILS = {e.strip().lower() for e in os.getenv("ADMIN_EMAILS", "aiseckiy@gmail.com").split(",")}
    if user.email in _ADMIN_EMAILS and user.role != "admin":
        user.role = "admin"
        db.commit()
    resp = JSONResponse({"ok": True, "name": user.name, "role": user.role})
    resp.set_cookie("lunary_session", _make_user_session(user), httponly=True, samesite="lax", max_age=60*60*24*30)
    return resp


# ─── Google OAuth ─────────────────────────────────────────────
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI", "https://www.lunary.kz/auth/google/callback")

@app.get("/auth/google")
def google_auth():
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=500, detail="Google OAuth не настроен")
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": GOOGLE_REDIRECT_URI,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "offline",
    }
    url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
    return RedirectResponse(url)


@app.get("/auth/google/callback")
def google_callback(code: str, request: Request, db: Session = Depends(get_db)):
    import hashlib
    from database import User as UserModel

    # Обмен code на токен
    token_data = urllib.parse.urlencode({
        "code": code,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "redirect_uri": GOOGLE_REDIRECT_URI,
        "grant_type": "authorization_code",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=token_data)
    try:
        with urllib.request.urlopen(req) as r:
            token = json.loads(r.read())
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка токена: {e}")

    # Получить профиль пользователя
    access_token = token.get("access_token")
    req2 = urllib.request.Request(
        "https://www.googleapis.com/oauth2/v2/userinfo",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    with urllib.request.urlopen(req2) as r:
        profile = json.loads(r.read())

    email = profile.get("email", "")
    google_id = profile.get("id", "")
    name = profile.get("name", "")
    avatar = profile.get("picture", "")

    _ADMIN_EMAILS = {e.strip().lower() for e in os.getenv("ADMIN_EMAILS", "aiseckiy@gmail.com").split(",")}
    auto_role = "admin" if email.lower() in _ADMIN_EMAILS else "customer"

    # Найти или создать пользователя
    user = db.query(UserModel).filter(UserModel.email == email).first()
    if not user:
        user = UserModel(email=email, google_id=google_id, name=name, avatar=avatar, role=auto_role)
        db.add(user)
        db.commit()
        db.refresh(user)
    else:
        user.google_id = google_id
        user.name = name
        user.avatar = avatar
        if email.lower() in _ADMIN_EMAILS and user.role != "admin":
            user.role = "admin"
        db.commit()

    next_url = "/admin" if user.role in ("admin", "manager") else "/shop"
    resp = RedirectResponse(next_url, status_code=302)
    resp.set_cookie("lunary_session", _make_user_session(user), httponly=True, samesite="lax", max_age=60*60*24*30)
    return resp


@app.get("/api/auth/check")
def auth_check():
    has_password = bool(os.getenv("ADMIN_PASSWORD", ""))
    return {"required": has_password}


@app.get("/api/auth/me")
def auth_me(request: Request):
    user = _get_user_from_session(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    return user


@app.patch("/api/auth/profile")
def update_profile(data: dict, request: Request, db: Session = Depends(get_db)):
    import hashlib
    from database import User as UserModel
    user = _get_user_from_session(request)
    if not user or not user.get("id"):
        raise HTTPException(status_code=401)
    u = db.query(UserModel).filter(UserModel.id == user["id"]).first()
    if not u:
        raise HTTPException(status_code=404)
    if "name" in data and data["name"]:
        u.name = data["name"].strip()
    if "phone" in data:
        u.phone = data["phone"].strip()
    if "password" in data and data["password"]:
        if len(data["password"]) < 6:
            raise HTTPException(status_code=400, detail="Минимум 6 символов")
        u.password_hash = hashlib.sha256(data["password"].encode()).hexdigest()
    db.commit()
    return {"ok": True}




# ─── Товары ──────────────────────────────────────────────────

# ─── Публичный магазин ───────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def root_page():
    return RedirectResponse("/shop", status_code=302)


from helpers import parse_images as _parse_images  # noqa: E402











# ─── Kaspi заказы ────────────────────────────────────────────




class KaspiOrdersPayload(BaseModel):
    orders: list


from helpers import (  # noqa: E402
    ARCHIVE_STATES, DEDUCT_STATES, CANCEL_STATES,
    send_tg_notification as _send_tg_notification,
    format_order_notification as _format_order_notification,
    find_product_by_sku as _find_product_by_sku,
    deduct_stock_for_order as _deduct_stock_for_order,
    return_stock_for_order as _return_stock_for_order,
)


@app.post("/api/kaspi/orders/sync")
def kaspi_orders_sync(payload: KaspiOrdersPayload, db: Session = Depends(get_db)):
    """Принимает заказы от локального sync скрипта и сохраняет в БД"""
    orders = payload.orders
    from database import KaspiOrder
    from helpers import build_sku_index
    added = 0
    updated = 0
    deducted_total = 0
    new_orders = []
    # Дедупликация входящих заказов
    seen = {}
    for o in orders:
        oid = _decode_kaspi_order_id(o["id"])
        seen[oid] = o
    orders = list(seen.values())

    # Индекс SKU→product_id — один раз на batch
    sku_index = build_sku_index(db)

    for o in orders:
        oid = _decode_kaspi_order_id(o["id"])
        existing = db.query(KaspiOrder).filter(KaspiOrder.order_id == oid).first()
        if not existing and oid != str(o["id"]):
            existing = db.query(KaspiOrder).filter(KaspiOrder.order_id == str(o["id"])).first()
            if existing:
                existing.order_id = oid
        new_state = o.get("state", "")
        if existing:
            old_state = existing.state
            existing.state = new_state
            updated += 1
            # Списываем при переходе в доставку/архив
            if new_state in DEDUCT_STATES and old_state not in DEDUCT_STATES and not existing.stock_deducted:
                _deduct_stock_for_order(existing, db, sku_index)
                existing.stock_deducted = 1
                deducted_total += 1
            # Возвращаем при отмене
            elif new_state in CANCEL_STATES and old_state not in CANCEL_STATES and existing.stock_deducted:
                _return_stock_for_order(existing, db, sku_index)
                existing.stock_deducted = 0
        else:
            ko = KaspiOrder(
                order_id=oid,
                state=new_state,
                total=int(o.get("total", 0)),
                customer=o.get("customer", ""),
                entries=json.dumps(o.get("entries", []), ensure_ascii=False),
                order_date=str(o.get("date", "")),
                stock_deducted=0,
            )
            db.add(ko)
            db.flush()
            added += 1
            new_orders.append(o)
            # Если новый заказ сразу в доставке или архиве — списываем
            if new_state in DEDUCT_STATES:
                _deduct_stock_for_order(ko, db, sku_index)
                ko.stock_deducted = 1
                deducted_total += 1
    db.commit()

    # Уведомления о новых заказах (только ACCEPTED/PICKUP/KASPI_DELIVERY)
    notify_states = {"ACCEPTED", "PICKUP", "KASPI_DELIVERY"}
    for o in new_orders:
        if o.get("state") in notify_states:
            _send_tg_notification(_format_order_notification(o))

    return {"added": added, "updated": updated, "stock_deducted": deducted_total}







# ── Состояние фонового импорта истории ───────────────────────
_history_import_state: dict = {
    "running": False, "done": False, "created": 0, "skipped": 0,
    "errors": 0, "chunk": 0, "total_chunks": 0, "log": []
}

def _run_history_import_bg():
    """Фоновый импорт заказов Kaspi за последний год с полным составом"""
    import time as _time
    from database import KaspiOrder, SessionLocal as _SL
    from sqlalchemy import func

    global _history_import_state
    s = _history_import_state

    STATES = ["COMPLETED", "ARCHIVE", "CANCELLED", "RETURN", "ACCEPTED", "NEW"]
    CHUNK_DAYS = 14
    TOTAL_DAYS = 365

    now_ms = int(_time.time() * 1000)
    day_ms = 24 * 60 * 60 * 1000
    chunks = list(range(0, TOTAL_DAYS, CHUNK_DAYS))
    with _state_lock:
        s["total_chunks"] = len(chunks) * len(STATES)
        s["chunk"] = 0

    db = _SL()
    try:
        for chunk_start in chunks:
            date_le = now_ms - chunk_start * day_ms
            date_ge = now_ms - (chunk_start + CHUNK_DAYS) * day_ms

            for state in STATES:
                if not s["running"]:
                    return
                s["chunk"] += 1
                page = 0
                chunk_created = 0

                while True:
                    try:
                        data = kaspi_module._proxy("get_orders", {
                            "state": state, "page": page, "size": 100,
                            "creationDateGe": str(date_ge),
                            "creationDateLe": str(date_le),
                        })
                    except Exception as e:
                        s["errors"] += 1
                        break

                    if not data:
                        break

                    items = data.get("data", [])
                    if not items:
                        break

                    for item in items:
                        attr = item.get("attributes", {})
                        oid = item.get("id", "")
                        if not oid:
                            continue

                        exists = db.query(KaspiOrder).filter(KaspiOrder.order_id == oid).first()
                        if exists:
                            s["skipped"] += 1
                            continue

                        # Получаем состав заказа (product_name, quantity)
                        pname, qty, sku_val = "", 1, ""
                        try:
                            entries_data = kaspi_module._proxy("get_order_entries", {"orderId": oid})
                            if entries_data and entries_data.get("data"):
                                entry = entries_data["data"][0]
                                ea = entry.get("attributes", {})
                                qty = int(ea.get("quantity", 1))
                                # Название из included
                                included = {i["id"]: i for i in entries_data.get("included", [])}
                                rels = entry.get("relationships", {})
                                prod_rel = rels.get("product", {}).get("data") or rels.get("offer", {}).get("data")
                                if prod_rel:
                                    prod = included.get(prod_rel.get("id"), {})
                                    pname = prod.get("attributes", {}).get("name", "")
                                    sku_val = ea.get("merchantSku", "")
                                if not pname:
                                    pname = ea.get("name", "") or ea.get("merchantSku", "")
                        except Exception:
                            pass

                        # Дата заказа
                        creation_ms = attr.get("creationDate", 0)
                        try:
                            import datetime as _dt
                            order_date = _dt.datetime.fromtimestamp(int(creation_ms) / 1000).strftime("%d.%m.%Y")
                        except Exception:
                            order_date = ""

                        total = int(attr.get("totalPrice", 0) or 0)

                        o = KaspiOrder(
                            order_id=oid,
                            state=attr.get("state", state),
                            total=total,
                            order_date=order_date,
                            product_name=pname or None,
                            sku=sku_val or None,
                            quantity=qty,
                            stock_deducted=0,
                        )
                        db.add(o)
                        s["created"] += 1
                        chunk_created += 1

                    db.commit()

                    if len(items) < 100:
                        break
                    page += 1
                    _time.sleep(0.2)

                if chunk_created > 0:
                    import datetime as _dt2
                    d_from = _dt2.datetime.fromtimestamp(date_ge / 1000).strftime("%d.%m")
                    d_to   = _dt2.datetime.fromtimestamp(date_le / 1000).strftime("%d.%m")
                    s["log"].append(f"{state} {d_from}–{d_to}: +{chunk_created}")
                    if len(s["log"]) > 30:
                        s["log"] = s["log"][-30:]

                _time.sleep(0.3)

    finally:
        db.close()
        with _state_lock:
            s["running"] = False
            s["done"] = True


@app.post("/api/kaspi/import-history")
def kaspi_import_history(request: Request):
    """Запустить фоновый исторический импорт заказов Kaspi за год"""
    user = _get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403)

    global _history_import_state
    with _state_lock:
        if _history_import_state["running"]:
            return {"ok": False, "detail": "Импорт уже запущен"}
        # Сбрасываем поля на месте, а не создаём новый dict — чтобы
        # bg thread видел тот же объект (он хранит ссылку в s).
        _history_import_state.update({
            "running": True, "done": False, "created": 0, "skipped": 0,
            "errors": 0, "chunk": 0, "total_chunks": 0, "log": []
        })

    t = threading.Thread(target=_run_history_import_bg, daemon=True)
    t.start()
    return {"ok": True, "detail": "Импорт запущен в фоне"}


@app.get("/api/kaspi/import-history/status")
def kaspi_import_history_status(request: Request):
    """Статус фонового импорта истории"""
    user = _get_user_from_session(request)
    if not user or not _is_staff(user):
        raise HTTPException(status_code=403)
    with _state_lock:
        # shallow copy под lock — делает snapshot для безопасного чтения
        snapshot = dict(_history_import_state)
        snapshot["log"] = list(snapshot.get("log", []))
    return snapshot


@app.post("/api/kaspi/import-history/stop")
def kaspi_import_history_stop(request: Request):
    """Остановить импорт"""
    user = _get_user_from_session(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403)
    with _state_lock:
        _history_import_state["running"] = False
    return {"ok": True}




@app.get("/api/products/export/xlsx")
def products_export_xlsx(db: Session = Depends(get_db), ids: str = None):
    """Экспорт остатков в Excel (формат загрузки Kaspi)"""
    from fastapi.responses import StreamingResponse
    from sqlalchemy import func
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from database import Product, Movement

    q = (
        db.query(Product, func.coalesce(func.sum(Movement.quantity), 0).label("stock"))
        .outerjoin(Movement, Movement.product_id == Product.id)
        .filter(Product.category != "Накладные")
        .group_by(Product.id)
        .order_by(Product.name)
    )
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            q = q.filter(Product.id.in_(id_list))
    rows = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остатки"

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="111827")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = ["SKU (Kaspi)", "Артикул (внутр.)", "Название товара", "Бренд",
               "Категория", "Цена (₸)", "Закуп (₸)", "Маржа %", "Остаток", "Мин. остаток", "Ед. изм.", "Поставщик"]
    col_widths = [28, 18, 50, 15, 20, 12, 12, 10, 12, 14, 10, 30]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    low_fill = PatternFill("solid", fgColor="FEE2E2")
    ok_fill  = PatternFill("solid", fgColor="DCFCE7")

    for ri, (p, stock) in enumerate(rows, 2):
        stock = int(stock)
        values = [
            p.kaspi_sku or "",
            p.kaspi_sku or "",
            p.name or "",
            p.brand or "",
            p.category or "",
            p.price or "",
            p.cost_price or "",
            round((p.price - p.cost_price) / p.price * 100) if p.price and p.cost_price and p.price > 0 else "",
            stock,
            p.min_stock or 0,
            p.unit or "шт",
            p.supplier or "",
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(ri, ci, val)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin

        # Подсветить строку если остаток низкий
        if stock <= (p.min_stock or 5):
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = low_fill
        ws.row_dimensions[ri].height = 18

    # Заморозить шапку
    ws.freeze_panes = "A2"

    # Второй лист — только для загрузки в Kaspi (формат active.xlsx)
    ws2 = wb.create_sheet("Для Kaspi")
    kaspi_headers = ["SKU", "model", "brand", "price", "PP1", "PP2", "PP3", "PP4", "PP5", "preorder"]
    for ci, h in enumerate(kaspi_headers, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 15

    for ri, (p, stock) in enumerate(rows, 2):
        if not p.kaspi_sku:
            continue
        ws2.cell(ri, 1, p.kaspi_sku)
        ws2.cell(ri, 2, p.name)
        ws2.cell(ri, 3, p.brand or "")
        ws2.cell(ri, 4, p.price or "")
        ws2.cell(ri, 5, max(0, int(stock)))  # PP1 = основной склад
        for c in range(6, 11):
            ws2.cell(ri, c, "no")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock_export.xlsx"}
    )


@app.get("/api/admin/orders-debug")
def orders_debug(request: Request, db: Session = Depends(get_db)):
    """Статистика заказов по статусам — для диагностики"""
    user = _get_user_from_session(request)
    if not user or user["role"] != "admin":
        raise HTTPException(status_code=403)
    from database import KaspiOrder
    from sqlalchemy import func
    rows = db.query(KaspiOrder.state, func.count(KaspiOrder.id), func.sum(KaspiOrder.total))\
        .group_by(KaspiOrder.state).all()
    result = []
    for state, cnt, total in sorted(rows, key=lambda x: x[1], reverse=True):
        result.append({"state": state, "count": cnt, "total": int(total or 0)})
    grand_total = sum(r["total"] for r in result)
    completed_total = sum(r["total"] for r in result if r["state"] in ("ARCHIVE", "Выдан"))
    return {"by_state": result, "grand_total": grand_total, "completed_total": completed_total}







@app.get("/api/purchases/list")
def purchases_list(db: Session = Depends(get_db)):
    """Список товаров с остатком ниже минимума, сгруппированных по поставщику."""
    all_stocks = crud.get_all_stocks(db)
    result = []
    for s in all_stocks:
        p = s["product"]
        stock = s["stock"]
        min_s = p.min_stock or 0
        if stock <= min_s:
            result.append({
                "id": p.id,
                "name": p.name,
                "sku": p.kaspi_sku or "",
                "supplier": p.supplier or "Не указан",
                "brand": p.brand or "",
                "stock": stock,
                "min_stock": min_s,
                "unit": p.unit or "шт",
                "need": max(1, min_s - stock + max(min_s, 5)),
            })
    # Сортируем: сначала по поставщику, потом по бренду
    result.sort(key=lambda x: (x["supplier"], x["brand"], x["name"]))
    return {"items": result, "total": len(result)}


@app.post("/api/purchases/send-tg")
def purchases_send_tg(db: Session = Depends(get_db)):
    """Отправить список закупок в Telegram."""
    data = purchases_list(db)
    items = data["items"]
    if not items:
        return {"ok": True, "message": "Нет товаров для закупки"}
    # Группируем по поставщику
    from collections import defaultdict
    by_supplier = defaultdict(list)
    for item in items:
        by_supplier[item["supplier"]].append(item)
    lines = ["📦 <b>Список закупок</b>", f"Итого позиций: {len(items)}", ""]
    for supplier, sup_items in by_supplier.items():
        lines.append(f"🏭 <b>{supplier}</b>")
        for it in sup_items:
            lines.append(f"  • {it['name']} ({it['sku']}) — осталось {it['stock']} {it['unit']}, заказать ~{it['need']}")
        lines.append("")
    _send_tg_notification("\n".join(lines))
    return {"ok": True, "message": f"Отправлено {len(items)} позиций"}


@app.post("/api/import-products")
def import_products(db: Session = Depends(get_db)):
    """Одноразовый импорт товаров из export_products.json"""
    import os
    path = os.path.join(os.path.dirname(__file__), '_archive', 'export_products.json')
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="export_products.json не найден")
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    from database import Product as _P
    added = skipped = 0
    for p in data:
        sku = p['sku'].upper()
        if False:  # sku field removed
            skipped += 1
            continue
        new_p = crud.create_product(
            name=p['name'], sku=sku, db=db,
            barcode=p.get('barcode'), category=p.get('category', 'Общее'),
            unit=p.get('unit', 'шт'), min_stock=p.get('min_stock', 5),
            brand=p.get('brand')
        )
        if p.get('stock', 0) > 0:
            crud.set_initial_stock(new_p.id, p['stock'], db)
        added += 1
    return {"added": added, "skipped": skipped}


@app.post("/api/kaspi/import-xml-products")
async def kaspi_import_xml_products(
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """UPSERT-импорт Kaspi каталога из XML.

    Для каждого <offer>:
    - Если товар уже есть (по kaspi_sku) — обновляем только name/brand/price.
      Всё остальное НЕ трогаем.
    - Если новый — создаём с category="Kaspi".
    - Stock синхронизируется через корректирующее движение (slaves пропускаются).
    """
    import xml.etree.ElementTree as ET
    import re
    import traceback
    from database import Product as _P, Movement

    try:
        if file:
            content = await file.read()
            original_name = file.filename or "ACTIVE.xml"
        else:
            path = os.path.join(os.path.dirname(__file__), 'ACTIVE.xml')
            if not os.path.exists(path):
                raise HTTPException(status_code=404, detail="ACTIVE.xml не найден. Загрузите файл.")
            with open(path, "rb") as f:
                content = f.read()
            original_name = "ACTIVE.xml"

        try:
            root = ET.fromstring(content)
        except ET.ParseError as pe:
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга XML: {pe}")

        def tag(el):
            return re.sub(r'\{[^}]+\}', '', el.tag)

        def find_text(el, name):
            for child in el:
                if tag(child) == name:
                    return (child.text or "").strip()
            return ""

        offers_el = None
        for child in root:
            if tag(child) == "offers":
                offers_el = child
                break

        if offers_el is None:
            raise HTTPException(status_code=400, detail="Тег <offers> не найден в XML")

        # Индекс существующих Kaspi-товаров по kaspi_sku (один SELECT)
        existing_by_sku: dict = {}
        for p in db.query(_P).filter(_P.kaspi_sku.isnot(None)).all():
            for s in (p.kaspi_sku or "").split(","):
                s = s.strip()
                if s and s not in existing_by_sku:
                    existing_by_sku[s] = p

        # Текущие остатки одним запросом
        from sqlalchemy import func
        stock_rows = db.query(Movement.product_id, func.coalesce(func.sum(Movement.quantity), 0)).group_by(Movement.product_id).all()
        current_stock_map = {pid: int(s) for pid, s in stock_rows}

        created = 0
        updated = 0
        unchanged = 0
        stock_adjusted = 0
        price_changed = 0

        for offer in offers_el:
            if tag(offer) != "offer":
                continue
            kaspi_sku = offer.attrib.get("sku", "").strip()
            if not kaspi_sku:
                continue

            model = find_text(offer, "model") or kaspi_sku
            brand = find_text(offer, "brand") or None

            price = None
            stock_count = 0
            for child in offer:
                t = tag(child)
                if t == "cityprices":
                    for cp in child:
                        if tag(cp) == "cityprice":
                            try:
                                price = int(cp.text.strip())
                            except Exception:
                                pass
                            break
                elif t == "availabilities":
                    for av in child:
                        if tag(av) == "availability":
                            try:
                                stock_count = int(float(av.attrib.get("stockCount", 0)))
                            except Exception:
                                pass
                            break

            existing = existing_by_sku.get(kaspi_sku)

            if existing:
                # UPDATE: обновляем только то что пришло из XML. Остальное НЕ трогаем.
                changed = False
                if model and existing.name != model:
                    existing.name = model
                    changed = True
                if brand and existing.brand != brand:
                    existing.brand = brand
                    changed = True
                if price is not None and existing.price != price:
                    existing.price = price
                    price_changed += 1
                    changed = True

                # Stock-коррекция. Для slave'ов в link-группе пропускаем
                # (иначе дельта N раз сбросится в мастера).
                delta = 0
                is_slave = bool(getattr(existing, "link_master_id", None))
                if not is_slave:
                    cur_stock = current_stock_map.get(existing.id, 0)
                    delta = stock_count - cur_stock
                    if delta != 0:
                        move_type = "income" if delta > 0 else "writeoff"
                        crud.add_movement(
                            existing.id, abs(delta), move_type, db,
                            source="kaspi_xml_import",
                            note=f"Коррекция остатка из Kaspi XML: было {cur_stock}, стало {stock_count}",
                        )
                        stock_adjusted += 1

                if changed or delta != 0:
                    updated += 1
                else:
                    unchanged += 1
            else:
                # CREATE: новый товар
                new_p = _P(
                    name=model,
                    kaspi_sku=kaspi_sku,
                    brand=brand,
                    price=price,
                    category="Kaspi",
                    unit="шт",
                    min_stock=1,
                )
                db.add(new_p)
                db.flush()
                if stock_count > 0:
                    crud.set_initial_stock(new_p.id, stock_count, db)
                created += 1

        db.commit()

        # Чистка мульти-SKU после импорта
        try:
            from sqlalchemy import text as _text
            sku_cleaned = db.execute(_text(
                "UPDATE products SET kaspi_sku = TRIM(SPLIT_PART(kaspi_sku, ',', 1)) WHERE kaspi_sku LIKE '%,%'"
            ))
            db.commit()
            sku_cleaned_count = sku_cleaned.rowcount
        except Exception:
            sku_cleaned_count = 0

        try:
            _save_upload(content, original_name, "kaspi_active", created + updated, db)
        except Exception as ue:
            print(f"[import-xml] save_upload warning: {ue}", flush=True)
        return {
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "stock_adjusted": stock_adjusted,
            "price_changed": price_changed,
            "sku_cleaned": sku_cleaned_count,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        tb = traceback.format_exc()
        print(f"[import-xml] FAILED: {e}\n{tb}", flush=True)
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:400]}")


@app.post("/api/kaspi/import-archive")
async def kaspi_import_archive(
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """UPSERT-импорт архивных товаров Kaspi.

    - Новые товары (отсутствующие в БД) — создаются.
    - Существующие — обновляются name/brand/price если что-то изменилось.
    - Stock НЕ трогается: в архиве stockCount=0, но физически товар может
      ещё лежать на складе. Остатки обновит следующий импорт ACTIVE.xml
      или Kaspi sync loop.
    """
    import xml.etree.ElementTree as ET
    import re
    import traceback
    from database import Product as _P

    try:
        if file:
            content = await file.read()
            original_name = file.filename or "ARCHIVE.xml"
        else:
            path = os.path.join(os.path.dirname(__file__), 'ARCHIVE.xml')
            if not os.path.exists(path):
                raise HTTPException(status_code=404, detail="ARCHIVE.xml не найден. Загрузите файл.")
            with open(path, "rb") as f:
                content = f.read()
            original_name = "ARCHIVE.xml"

        try:
            root = ET.fromstring(content)
        except ET.ParseError as pe:
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга XML: {pe}")

        def tag(el):
            return re.sub(r'\{[^}]+\}', '', el.tag)

        def find_text(el, name):
            for child in el:
                if tag(child) == name:
                    return (child.text or "").strip()
            return ""

        # Индекс существующих Kaspi-товаров по kaspi_sku
        existing_by_sku: dict = {}
        for p in db.query(_P).filter(_P.kaspi_sku.isnot(None)).all():
            for s in (p.kaspi_sku or "").split(","):
                s = s.strip()
                if s and s not in existing_by_sku:
                    existing_by_sku[s] = p

        offers_el = None
        for child in root:
            if tag(child) == "offers":
                offers_el = child
                break

        if offers_el is None:
            raise HTTPException(status_code=400, detail="Тег <offers> не найден в XML")

        created = 0
        updated = 0
        unchanged = 0
        price_changed = 0

        for offer in offers_el:
            if tag(offer) != "offer":
                continue
            kaspi_sku = offer.attrib.get("sku", "").strip()
            if not kaspi_sku:
                continue

            model = find_text(offer, "model") or kaspi_sku
            brand = find_text(offer, "brand") or None

            price = None
            for child in offer:
                if tag(child) == "cityprices":
                    for cp in child:
                        if tag(cp) == "cityprice":
                            try:
                                price = int(cp.text.strip())
                            except Exception:
                                pass
                            break
                    break

            existing = existing_by_sku.get(kaspi_sku)

            if existing:
                # UPDATE: обновляем name/brand/price. Stock и всё остальное НЕ трогаем.
                changed = False
                if model and existing.name != model:
                    existing.name = model
                    changed = True
                if brand and existing.brand != brand:
                    existing.brand = brand
                    changed = True
                if price is not None and existing.price != price:
                    existing.price = price
                    price_changed += 1
                    changed = True
                if changed:
                    updated += 1
                else:
                    unchanged += 1
            else:
                # CREATE: новый товар без начального stock (архивный = продан в Kaspi)
                new_p = _P(
                    name=model,
                    kaspi_sku=kaspi_sku,
                    brand=brand,
                    price=price,
                    category="Kaspi",
                    unit="шт",
                    min_stock=1,
                )
                db.add(new_p)
                existing_by_sku[kaspi_sku] = new_p
                created += 1

        db.commit()

        # Чистка мульти-SKU после импорта
        try:
            from sqlalchemy import text as _text
            sku_cleaned = db.execute(_text(
                "UPDATE products SET kaspi_sku = TRIM(SPLIT_PART(kaspi_sku, ',', 1)) WHERE kaspi_sku LIKE '%,%'"
            ))
            db.commit()
            sku_cleaned_count = sku_cleaned.rowcount
        except Exception:
            sku_cleaned_count = 0

        try:
            _save_upload(content, original_name, "kaspi_archive", created + updated, db)
        except Exception as ue:
            print(f"[import-archive] save_upload warning: {ue}", flush=True)
        return {
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "price_changed": price_changed,
            # Совместимость со старым фронтом
            "added": created,
            "skipped": unchanged,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        tb = traceback.format_exc()
        print(f"[import-archive] FAILED: {e}\n{tb}", flush=True)
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:400]}")


# ── Слияние накладных и Kaspi товаров ────────────────────────────────────────



@app.get("/import")
def import_page():
    from fastapi.responses import FileResponse
    return FileResponse("static/import.html")


# ── Справочник накладных (price_list_items) ───────────────────────────────────


# ── Загруженные файлы ─────────────────────────────────────────────────────────

@app.post("/api/fill-articles-from-pricelist")
def fill_articles_from_pricelist(db: Session = Depends(get_db)):
    """Заполняет barcode/kaspi_article у всех товаров у которых они пустые,
    матчинг по нечёткому имени из products_info.xml"""
    import xml.etree.ElementTree as ET
    import re
    from database import Product as _P

    xml_path = "products_info.xml"
    try:
        tree = ET.parse(xml_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось открыть XML: {e}")

    root = tree.getroot()
    товары = root.find("Товары")
    if товары is None:
        raise HTTPException(status_code=400, detail="Нет тега <Товары> в XML")

    STOPWORDS = {"и","в","на","с","для","из","по","шт","мл","л","кг","г","см","мм","м","гр","х","x","the","of","for","pcs"}
    def words(text):
        return {w for w in re.split(r'[\s\-_/,.()\[\]"«»\']+', (text or "").lower()) if len(w) > 2 and w not in STOPWORDS}

    # Загружаем прайс в память
    price_items = []
    for товар in товары.findall("Товар"):
        name_el = товар.find("Наименование")
        article_el = товар.find("Артикул")
        supplier_el = товар.find("Поставщик")
        price_el = товар.find("ЦенаЗакупки")
        name = (name_el.text or "").strip() if name_el is not None else ""
        article = (article_el.text or "").strip() if article_el is not None else ""
        supplier = (supplier_el.text or "").strip() if supplier_el is not None else ""
        cost_price_raw = (price_el.text or "0").strip() if price_el is not None else "0"
        if not name:
            continue
        try:
            cost_price = int(float(cost_price_raw))
        except Exception:
            cost_price = None
        price_items.append({"name": name, "article": article, "supplier": supplier,
                            "cost_price": cost_price, "words": words(name)})

    # Берём все товары у которых нет артикула
    all_products = db.query(_P).all()
    updated = 0

    for p in all_products:
        if p.barcode and p.supplier_article:
            continue  # артикулы уже есть

        p_words = words(p.name)
        if not p_words:
            continue

        best_score = 0.0
        best_item = None
        for item in price_items:
            if not item["words"]:
                continue
            common = p_words & item["words"]
            score = len(common) / max(len(p_words), len(item["words"]))
            if score > best_score:
                best_score = score
                best_item = item

        if best_item and best_score >= 0.5:
            changed = False
            if best_item["article"] and not p.barcode:
                p.barcode = best_item["article"]
                changed = True
            if best_item["article"] and not p.supplier_article:
                p.supplier_article = best_item["article"]
                changed = True
            if best_item["cost_price"] is not None and not p.cost_price:
                p.cost_price = best_item["cost_price"]
                changed = True
            if best_item["supplier"] and not p.supplier:
                p.supplier = best_item["supplier"]
                changed = True
            if changed:
                updated += 1

    db.commit()
    return {"updated": updated, "total_products": len(all_products)}


@app.post("/api/import-price-list")
async def import_price_list(
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """Импортирует закупочные цены из XML прайс-листа в таблицу products."""
    import xml.etree.ElementTree as ET
    import re
    from database import Product as _P

    if file:
        content = await file.read()
        original_name = file.filename or "price_list.xml"
        root = ET.fromstring(content)
        товары = root.find("Товары")
    else:
        xml_path = "products_info.xml"
        original_name = "products_info.xml"
        try:
            with open(xml_path, "rb") as f:
                content = f.read()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Не удалось открыть XML: {e}")
        root = ET.fromstring(content)
        товары = root.find("Товары")

    if товары is None:
        raise HTTPException(status_code=400, detail="Нет тега <Товары> в XML")

    import uuid
    created = 0

    # Индекс существующих артикулов чтобы не дублировать
    existing_articles = {p.supplier_article for p in db.query(_P).filter(_P.supplier_article != None).all()}
    existing_kaspi_skus = {p.kaspi_sku for p in db.query(_P).filter(_P.kaspi_sku != None).all()}

    for товар in товары.findall("Товар"):
        name_el = товар.find("Наименование")
        price_el = товар.find("ЦенаЗакупки")
        supplier_el = товар.find("Поставщик")
        article_el = товар.find("Артикул")
        unit_el = товар.find("ЕдИзм")

        name = (name_el.text or "").strip() if name_el is not None else ""
        cost_price_raw = (price_el.text or "0").strip() if price_el is not None else "0"
        supplier = (supplier_el.text or "").strip() if supplier_el is not None else ""
        article = (article_el.text or "").strip() if article_el is not None else ""
        unit = (unit_el.text or "шт").strip() if unit_el is not None else "шт"

        if not name:
            continue

        try:
            cost_price = int(float(cost_price_raw))
        except Exception:
            cost_price = None

        # Не создаём дубль если артикул уже есть в базе
        if article and article in existing_articles:
            continue

        # Уникальный SKU
        sku = article.upper() if article else f"PL-{uuid.uuid4().hex[:8].upper()}"
        while sku in existing_skus:
            sku = f"PL-{uuid.uuid4().hex[:8].upper()}"

        new_p = _P(
            name=name,
            sku=sku,
            supplier_article=article if article else None,
            category="Накладные",
            unit=unit,
            cost_price=cost_price,
            supplier=supplier,
        )
        db.add(new_p)
        existing_kaspi_skus.add(sku)
        if article:
            existing_articles.add(article)
        created += 1

    db.commit()
    _save_upload(content, original_name, "price_list", created, db)
    return {"created": created}



# ── Дизайн-токены (тема) ─────────────────────────────────────────────────────

@app.get("/api/admin/processes")
def get_processes(db: Session = Depends(get_db)):
    """Статус фоновых процессов и БД."""
    from database import Product, KaspiOrder, Movement, PriceListItem, UploadedFile
    from sqlalchemy import func as sqlfunc

    uptime_sec = int(time.monotonic() - APP_START)
    h, rem = divmod(uptime_sec, 3600)
    m, s = divmod(rem, 60)
    uptime_str = f"{h}ч {m}м {s}с" if h else f"{m}м {s}с"

    # Счётчики БД
    db_counts = {}
    try:
        db_counts = {
            "products":   db.query(sqlfunc.count(Product.id)).scalar(),
            "kaspi":      db.query(sqlfunc.count(Product.id)).filter(Product.category == "Kaspi").scalar(),
            "orders":     db.query(sqlfunc.count(KaspiOrder.id)).scalar(),
            "movements":  db.query(sqlfunc.count(Movement.id)).scalar(),
            "pricelist":  db.query(sqlfunc.count(PriceListItem.id)).scalar(),
            "uploads":    db.query(sqlfunc.count(UploadedFile.id)).scalar(),
        }
        db_status = "ok"
    except Exception as e:
        db_status = f"error: {e}"

    # Kaspi sync — снимаем snapshot под lock чтобы не читать half-updated state
    with _state_lock:
        ks = dict(_PROCESS_STATUS["kaspi_sync"])
        server_start = _PROCESS_STATUS["server_start"]

    next_run_in = None
    if ks.get("next_run"):
        next_run_in = max(0, int(ks["next_run"] - datetime.utcnow().timestamp()))

    return {
        "uptime": uptime_str,
        "server_start": server_start,
        "db": {"status": db_status, **db_counts},
        "processes": [
            {
                "name": "Kaspi синхронизация",
                "key": "kaspi_sync",
                "status": ks["status"],
                "last_run": ks["last_run"],
                "next_run_in": next_run_in,
                "last_result": ks["last_result"],
                "last_error": ks["last_error"],
                "cycle_count": ks["cycle_count"],
                "interval": "5 мин",
            },
        ],
    }


