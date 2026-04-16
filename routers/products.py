"""Products CRUD + история движений + алерты низких остатков."""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func

from database import get_db
from helpers import get_user_from_session, is_admin
import crud

router = APIRouter(tags=["products"])


# ── Pydantic схемы ──────────────────────────────────────────
class ProductCreate(BaseModel):
    name: str
    sku: str
    barcode: Optional[str] = None
    category: str = "Общее"
    unit: str = "шт"
    min_stock: int = 5
    brand: Optional[str] = None
    price: Optional[int] = None


class ProductUpdate(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    min_stock: Optional[int] = None
    brand: Optional[str] = None
    price: Optional[int] = None
    kaspi_sku: Optional[str] = None
    kaspi_article: Optional[str] = None
    cost_price: Optional[int] = None
    supplier: Optional[str] = None
    supplier_article: Optional[str] = None
    description: Optional[str] = None
    specs: Optional[str] = None
    show_in_shop: Optional[bool] = None


class ProductPatch(BaseModel):
    barcode: Optional[str] = None


class SetStockBody(BaseModel):
    actual: int
    note: Optional[str] = None


class StockAdjust(BaseModel):
    quantity: int
    type: str
    source: str = "manual"
    note: Optional[str] = None


# ══════════════════════════════════════════════════════
# Endpoints
# ══════════════════════════════════════════════════════
@router.get("/api/products")
def list_products(db: Session = Depends(get_db)):
    from database import Product as _P
    from sqlalchemy import func as sqlfunc

    stocks = crud.get_all_stocks(db)

    # Индекс master_id → количество slaves (для бэйджика "N в группе" у мастера)
    group_counts = dict(
        db.query(_P.link_master_id, sqlfunc.count(_P.id))
        .filter(_P.link_master_id.isnot(None))
        .group_by(_P.link_master_id)
        .all()
    )

    result = []
    for s in stocks:
        p = s["product"]
        # group_size > 0 только у мастера (показывает количество slaves + сам мастер)
        group_size = group_counts.get(p.id, 0) + 1 if p.id in group_counts else 0
        result.append({
            "id": p.id,
            "name": p.name,
            "barcode": p.barcode,
            "category": p.category,
            "unit": p.unit,
            "min_stock": p.min_stock,
            "stock": s["stock"],      # для slaves уже подменён на stock мастера (см. crud.get_all_stocks)
            "low": s["stock"] <= p.min_stock,
            "brand": p.brand or "",
            "price": p.price,          # цена у каждого своя
            "kaspi_sku": p.kaspi_sku or "",
            "cost_price": p.cost_price,  # закуп у каждого свой
            "supplier": p.supplier or "",
            "supplier_article": p.supplier_article or "",
            "description": p.description or "",
            "specs": p.specs or "[]",
            "link_master_id": p.link_master_id,
            "link_group_size": group_size,  # >0 только у мастера
        })
    return result


@router.post("/api/products/import/xlsx")
async def products_import_xlsx(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Импорт товаров из Excel. Колонки: Название, Бренд, Категория, Цена, Закуп, SKU, Артикул, Ед. изм., Мин. остаток, Поставщик, Описание"""
    user = get_user_from_session(request)
    if not is_admin(user):
        raise HTTPException(status_code=403)

    import openpyxl, io
    from database import Product as _P

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не удалось открыть файл: {e}")

    ws = wb.active
    headers_row = [str(c.value or "").strip().lower() for c in ws[1]]

    COL_MAP = {
        "name":        ["название", "наименование", "товар", "name"],
        "brand":       ["бренд", "brand", "марка", "производитель"],
        "category":    ["категория", "category"],
        "price":       ["цена", "price", "цена (₸)", "розница"],
        "cost_price":  ["закуп", "себестоимость", "закупочная", "cost", "закуп (₸)"],
        "kaspi_sku":   ["sku", "kaspi sku", "sku (kaspi)", "артикул kaspi", "kaspi_sku"],
        "unit":        ["ед. изм.", "единица", "unit", "ед.изм", "ед изм"],
        "min_stock":   ["мин. остаток", "минимум", "min_stock", "мин остаток"],
        "supplier":    ["поставщик", "supplier"],
        "description": ["описание", "description"],
    }

    def find_col(field):
        for alias in COL_MAP[field]:
            for i, h in enumerate(headers_row):
                if alias in h or h in alias:
                    return i
        return None

    cols = {f: find_col(f) for f in COL_MAP}

    def val(row, field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx].value

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        name = val(row, "name")
        if not name or str(name).strip() == "":
            continue
        name = str(name).strip()

        try:
            price_raw = val(row, "price")
            price = float(str(price_raw).replace(" ", "").replace(",", ".")) if price_raw else None
            cost_raw = val(row, "cost_price")
            cost = float(str(cost_raw).replace(" ", "").replace(",", ".")) if cost_raw else None
            min_stock_raw = val(row, "min_stock")
            min_stock = int(float(str(min_stock_raw))) if min_stock_raw else 0
        except Exception:
            price = cost = None
            min_stock = 0

        kaspi_sku = str(val(row, "kaspi_sku") or "").strip() or None
        brand = str(val(row, "brand") or "").strip() or None
        category = str(val(row, "category") or "").strip() or None
        unit = str(val(row, "unit") or "").strip() or "шт"
        supplier = str(val(row, "supplier") or "").strip() or None
        description = str(val(row, "description") or "").strip() or None

        existing = None
        if kaspi_sku:
            existing = db.query(_P).filter(_P.kaspi_sku == kaspi_sku).first()
        if not existing:
            existing = db.query(_P).filter(_P.name == name).first()

        if existing:
            if brand: existing.brand = brand
            if category: existing.category = category
            if price: existing.price = price
            if cost: existing.cost_price = cost
            if kaspi_sku: existing.kaspi_sku = kaspi_sku
            if unit: existing.unit = unit
            if min_stock: existing.min_stock = min_stock
            if supplier: existing.supplier = supplier
            if description: existing.description = description
            updated += 1
        else:
            p = _P(
                name=name, brand=brand, category=category or "Другое",
                price=price, cost_price=cost, kaspi_sku=kaspi_sku,
                unit=unit, min_stock=min_stock, supplier=supplier,
                description=description,
            )
            db.add(p)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@router.get("/admin/import-xlsx", response_class=HTMLResponse)
def import_xlsx_page(request: Request):
    user = get_user_from_session(request)
    if not is_admin(user):
        return RedirectResponse("/login")
    with open("static/import_xlsx.html", encoding="utf-8") as f:
        return f.read()


@router.post("/api/products")
def create_product(data: ProductCreate, db: Session = Depends(get_db)):
    p = crud.create_product(
        name=data.name, db=db,
        barcode=data.barcode, category=data.category,
        unit=data.unit, min_stock=data.min_stock, brand=data.brand, price=data.price
    )
    return {"id": p.id, "name": p.name}


@router.get("/api/products/suppliers")
def get_suppliers(db: Session = Depends(get_db)):
    from database import Product as _P
    rows = db.query(_P.supplier).filter(
        _P.supplier.isnot(None), _P.supplier != "",
        _P.category != "Накладные"
    ).distinct().all()
    return {"suppliers": sorted([r[0] for r in rows if r[0]])}


@router.get("/api/products/stats")
def products_stats(db: Session = Depends(get_db)):
    from database import Product as _P
    rows = db.query(_P.category, func.count(_P.id)).filter(
        _P.category != "Накладные"
    ).group_by(_P.category).all()
    by_category = {cat: cnt for cat, cnt in rows}
    total = sum(by_category.values())
    return {"total": total, "by_category": by_category}


@router.get("/api/products/search")
def search_products(q: str, db: Session = Depends(get_db)):
    products = crud.find_product(q, db)
    if not products:
        return []
    ids = {p.id for p in products}
    all_stocks = crud.get_all_stocks(db)
    stock_map = {s["product"].id: s["stock"] for s in all_stocks if s["product"].id in ids}
    return [
        {
            "id": p.id,
            "name": p.name,
            "kaspi_sku": p.kaspi_sku or "",
            "barcode": p.barcode,
            "stock": stock_map.get(p.id, 0),
            "unit": p.unit,
            "cost_price": p.cost_price,
            "supplier": p.supplier or "",
        }
        for p in products
    ]


@router.get("/api/products/barcode/{barcode}")
def get_by_barcode(barcode: str, db: Session = Depends(get_db)):
    p = crud.get_product_by_barcode(barcode, db)
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    return {
        "id": p.id,
        "name": p.name,
        "sku": p.kaspi_sku or "",
        "barcode": p.barcode,
        "stock": crud.get_stock(p.id, db),
        "unit": p.unit,
        "min_stock": p.min_stock
    }


@router.delete("/api/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    from database import Product as _P, Movement as _M
    p = db.query(_P).filter(_P.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    # Если удаляем master link-группы — отвязываем всех slaves (становятся independent)
    db.query(_P).filter(_P.link_master_id == product_id).update(
        {"link_master_id": None}, synchronize_session=False
    )
    db.query(_M).filter(_M.product_id == product_id).delete()
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.put("/api/products/{product_id}")
def update_product(product_id: int, data: ProductUpdate, db: Session = Depends(get_db)):
    from database import Product as _P
    p = crud.get_product_by_id(product_id, db)
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates.pop("sku", None)
    if "barcode" in updates and updates["barcode"]:
        conflict = db.query(_P).filter(_P.barcode == updates["barcode"], _P.id != product_id).first()
        if conflict:
            raise HTTPException(status_code=409, detail=f"Штрихкод уже привязан к товару «{conflict.name}»")
    p = crud.update_product(product_id, db, **updates)
    return {"id": p.id, "name": p.name}


@router.get("/api/products/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db)):
    from database import Product as _P

    p = crud.get_product_by_id(product_id, db)
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    stock = crud.get_stock(product_id, db)  # для slave уже вернёт stock мастера

    # Link-группа: master отдаёт список slaves, slave — инфу о мастере.
    # Цены у каждого свои — не подменяем.
    master_info = None
    slaves_info = []

    if p.link_master_id:
        master = db.query(_P).filter(_P.id == p.link_master_id).first()
        if master:
            master_info = {"id": master.id, "name": master.name, "kaspi_sku": master.kaspi_sku or ""}
    else:
        slaves = db.query(_P).filter(_P.link_master_id == p.id).all()
        slaves_info = [
            {"id": s.id, "name": s.name, "kaspi_sku": s.kaspi_sku or "", "price": s.price}
            for s in slaves
        ]

    return {"product": {
        "id": p.id, "name": p.name, "kaspi_sku": p.kaspi_sku or "", "kaspi_article": p.kaspi_article or "",
        "category": p.category or "", "unit": p.unit or "шт",
        "price": p.price, "min_stock": p.min_stock,
        "stock": stock, "low": stock <= (p.min_stock or 0),
        "brand": p.brand or "", "supplier": p.supplier or "",
        "cost_price": p.cost_price, "barcode": p.barcode or "",
        "supplier_article": p.supplier_article or "",
        "images": p.images or "[]",
        "description": p.description or "",
        "specs": p.specs or "[]",
        "show_in_shop": bool(p.show_in_shop),
        "meta_title": p.meta_title or "",
        "meta_description": p.meta_description or "",
        "meta_keywords": p.meta_keywords or "",
        # Link-group info
        "link_master_id": p.link_master_id,
        "link_master": master_info,        # если slave: {id, name, kaspi_sku}
        "link_slaves": slaves_info,        # если master: список slaves
        "is_link_master": bool(slaves_info),
    }}


@router.patch("/api/products/{product_id}")
def patch_product(product_id: int, data: ProductPatch, db: Session = Depends(get_db)):
    from database import Product
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    if data.barcode is not None:
        if data.barcode:
            conflict = db.query(Product).filter(Product.barcode == data.barcode, Product.id != product_id).first()
            if conflict:
                raise HTTPException(
                    status_code=409,
                    detail=f"Штрихкод уже привязан к товару «{conflict.name}» (арт. {conflict.supplier_article or conflict.kaspi_sku or '?'})"
                )
        p.barcode = data.barcode
    db.commit()
    return {"ok": True}


@router.post("/api/products/{product_id}/set-stock")
def set_stock_value(product_id: int, data: SetStockBody, request: Request, db: Session = Depends(get_db)):
    p = crud.get_product_by_id(product_id, db)
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    if data.actual < 0:
        raise HTTPException(status_code=400, detail="Остаток не может быть отрицательным")
    current = crud.get_stock(product_id, db)
    delta = data.actual - current
    if delta == 0:
        return {"product": p.name, "new_stock": current, "delta": 0}
    move_type = "income" if delta > 0 else "writeoff"
    note = data.note or f"Коррекция остатка: было {current}, стало {data.actual}"
    u = get_user_from_session(request)
    crud.add_movement(product_id, abs(delta), move_type, db, "web", note,
                      user_id=u.get("id") if u else None,
                      user_name=u.get("name") or u.get("email") if u else None)
    new_stock = crud.get_stock(product_id, db)
    return {"product": p.name, "new_stock": new_stock, "delta": delta}


@router.get("/api/products/{product_id}/stock")
def get_stock_endpoint(product_id: int, db: Session = Depends(get_db)):
    p = crud.get_product_by_id(product_id, db)
    if not p:
        raise HTTPException(status_code=404, detail="Товар не найден")
    stock = crud.get_stock(product_id, db)
    return {"product_id": product_id, "name": p.name, "stock": stock, "unit": p.unit}


@router.post("/api/products/{product_id}/movement")
def add_movement_endpoint(product_id: int, data: StockAdjust, request: Request, db: Session = Depends(get_db)):
    try:
        p = crud.get_product_by_id(product_id, db)
        if not p:
            raise HTTPException(status_code=404, detail="Товар не найден")

        if data.type not in ("income", "sale", "writeoff", "return", "adjustment"):
            raise HTTPException(status_code=400, detail="Неверный тип движения")
        if data.quantity <= 0:
            raise HTTPException(status_code=400, detail="Количество должно быть больше нуля")

        u = get_user_from_session(request)
        m = crud.add_movement(product_id, data.quantity, data.type, db, data.source, data.note,
                              user_id=u.get("id") if u else None,
                              user_name=u.get("name") or u.get("email") if u else None)
        new_stock = crud.get_stock(product_id, db)
        return {
            "movement_id": m.id,
            "product": p.name,
            "type": data.type,
            "quantity": data.quantity,
            "new_stock": new_stock
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print("MOVEMENT ERROR:", traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/products/{product_id}/history")
def get_product_history(product_id: int, db: Session = Depends(get_db)):
    movements = crud.get_movements(product_id, db)
    p = crud.get_product_by_id(product_id, db)
    type_labels = {
        "income": "📦 Приход",
        "sale": "🛒 Продажа",
        "writeoff": "🗑 Списание",
        "return": "↩️ Возврат",
        "adjustment": "✏️ Корректировка"
    }
    return [
        {
            "id": m.id,
            "type": m.type,
            "type_label": type_labels.get(m.type, m.type),
            "quantity": m.quantity,
            "source": m.source,
            "note": m.note,
            "date": m.created_at.strftime("%d.%m.%Y %H:%M")
        }
        for m in movements
    ]


# ─── Журнал всех операций ────────────────────────────────────
@router.get("/api/history")
def get_all_history(
    limit: int = 100,
    offset: int = 0,
    type: Optional[str] = None,
    source: Optional[str] = None,
    product_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    from database import Movement, Product
    q = db.query(Movement, Product).join(Product, Movement.product_id == Product.id)
    if type:
        q = q.filter(Movement.type == type)
    if source:
        q = q.filter(Movement.source == source)
    if product_id:
        q = q.filter(Movement.product_id == product_id)

    total = q.count()
    rows = q.order_by(desc(Movement.created_at)).offset(offset).limit(limit).all()

    type_labels = {"income": "Приход", "sale": "Продажа", "writeoff": "Списание",
                   "return": "Возврат", "adjustment": "Корректировка"}
    return {
        "total": total,
        "items": [
            {
                "id": m.id,
                "product_id": p.id,
                "product_name": p.name,
                "product_sku": p.kaspi_sku or "",
                "type": m.type,
                "type_label": type_labels.get(m.type, m.type),
                "quantity": m.quantity,
                "source": m.source or "manual",
                "note": m.note or "",
                "date": m.created_at.strftime("%d.%m.%Y"),
                "time": m.created_at.strftime("%H:%M"),
                "datetime_iso": m.created_at.isoformat(),
                "user_name": m.user_name or "",
            }
            for m, p in rows
        ]
    }


@router.delete("/api/history/{movement_id}")
def delete_movement(movement_id: int, db: Session = Depends(get_db)):
    from database import Movement
    m = db.query(Movement).filter(Movement.id == movement_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    db.delete(m)
    db.commit()
    return {"ok": True}


# ─── Алерты ──────────────────────────────────────────────────
@router.get("/api/alerts/low-stock")
def low_stock(db: Session = Depends(get_db)):
    items = crud.get_low_stock_products(db)
    return [
        {"id": p.id, "name": p.name, "stock": stock, "min_stock": p.min_stock}
        for p, stock in items
    ]


# ─── Link-группы (общий stock для дубликатов карточек) ─────
class LinkGroupBody(BaseModel):
    master_id: int
    slave_ids: list  # list[int]


@router.post("/api/products/link")
def link_products(body: LinkGroupBody, request: Request, db: Session = Depends(get_db)):
    """Объединить товары в link-группу с общим остатком.

    master_id — главный товар (физически существующий на складе).
    slave_ids — зеркала (дубликаты карточек в Kaspi). Их текущие остатки
    переносятся мастеру через корректирующие движения (чтобы история stock'а
    нигде не терялась).
    """
    from database import Product as _P, Movement as _M
    from sqlalchemy import func as sqlfunc

    user = get_user_from_session(request)
    if not is_staff(user):
        raise HTTPException(status_code=403)

    if not body.slave_ids:
        raise HTTPException(status_code=400, detail="Не выбраны товары для привязки")
    if body.master_id in body.slave_ids:
        raise HTTPException(status_code=400, detail="Мастер не может быть slave самому себе")

    master = db.query(_P).filter(_P.id == body.master_id).first()
    if not master:
        raise HTTPException(status_code=404, detail="Мастер не найден")
    # Мастер не может сам быть slave'ом другой группы — иначе разбивать сначала
    if master.link_master_id:
        raise HTTPException(
            status_code=400,
            detail=f"Мастер уже привязан к другой группе (мастер id={master.link_master_id}). Сначала отвяжи его."
        )

    slaves = db.query(_P).filter(_P.id.in_(body.slave_ids)).all()
    if len(slaves) != len(body.slave_ids):
        raise HTTPException(status_code=400, detail="Один или несколько товаров не найдены")

    # Каждый slave отдаёт свои текущие остатки мастеру. Минус у slave
    # (чтобы на нём стало 0) и плюс у мастера (чтобы он получил все остатки).
    total_transferred = 0
    for s in slaves:
        if s.id == master.id:
            continue
        # Если slave сам был мастером группы — его slaves тоже переподвязываем к новому мастеру
        if db.query(_M).filter(False).first() is not None:
            pass  # no-op (оставлено чтоб не забыть; реальная обработка ниже)
        nested_slaves = db.query(_P).filter(_P.link_master_id == s.id).all()
        for ns in nested_slaves:
            ns.link_master_id = master.id

        # Считаем текущий stock у slave (читаем напрямую по movements — до того как привяжем)
        cur = db.query(sqlfunc.coalesce(sqlfunc.sum(_M.quantity), 0)).filter(
            _M.product_id == s.id
        ).scalar() or 0
        cur = int(cur)

        if cur != 0:
            # Снимаем со slave
            db.add(_M(
                product_id=s.id,
                quantity=-cur,
                type="adjustment",
                source="link_group",
                note=f"Объединение с мастером #{master.id}: остаток передан мастеру",
            ))
            # Зачисляем мастеру
            db.add(_M(
                product_id=master.id,
                quantity=cur,
                type="income",
                source="link_group",
                note=f"Получен остаток от слейва #{s.id} ({s.name[:40]}) при объединении",
            ))
            total_transferred += cur

        s.link_master_id = master.id

    db.commit()

    group_size = db.query(sqlfunc.count(_P.id)).filter(_P.link_master_id == master.id).scalar() or 0

    return {
        "ok": True,
        "master_id": master.id,
        "master_name": master.name,
        "slaves_linked": len(slaves),
        "group_size": group_size + 1,  # +1 = сам мастер
        "stock_transferred": total_transferred,
    }


@router.post("/api/products/unlink")
def unlink_products(body: dict, request: Request, db: Session = Depends(get_db)):
    """Отвязать товары от их master'ов. Остатки мастера НЕ перераспределяются —
    остаются на мастере. Если slave нужен с собственным остатком — руками
    сделай set-stock после unlink."""
    from database import Product as _P

    user = get_user_from_session(request)
    if not is_staff(user):
        raise HTTPException(status_code=403)

    ids = body.get("product_ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="Не выбраны товары")

    count = db.query(_P).filter(
        _P.id.in_(ids),
        _P.link_master_id.isnot(None)
    ).update({"link_master_id": None}, synchronize_session=False)
    db.commit()

    return {"ok": True, "unlinked": count}
